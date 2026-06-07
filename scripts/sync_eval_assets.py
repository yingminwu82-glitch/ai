"""
Sync evaluation nail assets from the customer Excel file.

Outputs:
- assets/styles/style-01..25.png
- assets/nail-tips-generated/tip-XX-{thumb,index,middle,ring,pinky}1.png
- assets/styles_manifest.json
"""
import json
import math
import os
import sys
import urllib.request
from pathlib import Path

import cv2
import mediapipe as mp
import numpy as np
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = (
    "/Users/yanghan/Library/Containers/com.tencent.xinWeChat/Data/Documents/"
    "xwechat_files/wxid_u35ma0htwoxh22_0002/temp/drag/"
    "命题三美甲评测数据（对外版）(1).xlsx"
)

STYLES_DIR = PROJECT_ROOT / "assets" / "styles"
TIPS_DIR = PROJECT_ROOT / "assets" / "nail-tips-generated"
MANIFEST_PATH = PROJECT_ROOT / "assets" / "styles_manifest.json"

FINGERS = {
    "thumb": (4, 3),
    "index": (8, 7),
    "middle": (12, 11),
    "ring": (16, 15),
    "pinky": (20, 19),
}


def download(url: str, dest: Path) -> bool:
    if dest.exists() and dest.stat().st_size > 1024:
        return True
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            data = resp.read()
        img_arr = np.frombuffer(data, dtype=np.uint8)
        img = cv2.imdecode(img_arr, cv2.IMREAD_COLOR)
        if img is None:
            print(f"skip invalid image: {url}")
            return False
        dest.parent.mkdir(parents=True, exist_ok=True)
        cv2.imwrite(str(dest), img)
        return True
    except Exception as exc:
        print(f"download failed: {url} -> {exc}")
        return False


def estimate_tip_crop(tip_xy, dip_xy, img_w, img_h, is_thumb=False):
    dx = dip_xy[0] - tip_xy[0]
    dy = dip_xy[1] - tip_xy[1]
    finger_len = math.hypot(dx, dy)
    if finger_len < 5:
        return None
    angle = math.atan2(dy, dx)
    offset = 0.55 if is_thumb else 0.35
    crop_w = finger_len * (1.1 if is_thumb else 1.0)
    crop_h = finger_len * (1.45 if is_thumb else 1.25)
    cx = tip_xy[0] + offset * finger_len * math.cos(angle)
    cy = tip_xy[1] + offset * finger_len * math.sin(angle)
    x0 = max(0, int(cx - crop_w / 2))
    y0 = max(0, int(cy - crop_h / 2))
    x1 = min(img_w, int(cx + crop_w / 2))
    y1 = min(img_h, int(cy + crop_h / 2))
    if x1 - x0 < 18 or y1 - y0 < 18:
        return None
    return x0, y0, x1, y1


def extract_tips(style_id: int, image_path: Path) -> int:
    img = cv2.imread(str(image_path), cv2.IMREAD_COLOR)
    if img is None:
        return 0
    h, w = img.shape[:2]
    rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    count = 0
    with mp.solutions.hands.Hands(static_image_mode=True, max_num_hands=2, min_detection_confidence=0.35) as hands:
        result = hands.process(rgb)
        if not result.multi_hand_landmarks:
            return 0
        # Prefer the first detected hand; style photos are expected to be single-hand showcase images.
        landmarks = result.multi_hand_landmarks[0].landmark
        pts = [(int(lm.x * w), int(lm.y * h)) for lm in landmarks]
        for finger, (tip_i, dip_i) in FINGERS.items():
            crop = estimate_tip_crop(pts[tip_i], pts[dip_i], w, h, is_thumb=finger == "thumb")
            if not crop:
                continue
            x0, y0, x1, y1 = crop
            tip = img[y0:y1, x0:x1]
            if tip.size == 0:
                continue
            out = cv2.resize(tip, (96, 120), interpolation=cv2.INTER_CUBIC)
            out_path = TIPS_DIR / f"tip-{style_id:02d}-{finger}1.png"
            cv2.imwrite(str(out_path), out)
            count += 1
    return count


def mean_hex(image_path: Path) -> str:
    img = cv2.imread(str(image_path), cv2.IMREAD_COLOR)
    if img is None:
        return "#d9b8ad"
    h, w = img.shape[:2]
    crop = img[int(h * 0.2):int(h * 0.8), int(w * 0.2):int(w * 0.8)]
    small = cv2.resize(crop, (1, 1), interpolation=cv2.INTER_AREA)[0, 0]
    b, g, r = [int(v) for v in small]
    return f"#{r:02x}{g:02x}{b:02x}"


def read_style_rows(xlsx_path: str):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["款式图"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid, original_url, enhanced_url = row[:3]
        if not sid or not enhanced_url:
            continue
        rows.append({
            "id": int(sid),
            "original_url": str(original_url or ""),
            "enhanced_url": str(enhanced_url),
        })
    return rows


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    STYLES_DIR.mkdir(parents=True, exist_ok=True)
    TIPS_DIR.mkdir(parents=True, exist_ok=True)

    styles = []
    for row in read_style_rows(xlsx_path):
        sid = row["id"]
        style_path = STYLES_DIR / f"style-{sid:02d}.png"
        ok = download(row["enhanced_url"], style_path)
        tip_count = extract_tips(sid, style_path) if ok else 0
        styles.append({
            "id": sid,
            "name": f"训练款式 {sid:02d}",
            "color_hex": mean_hex(style_path),
            "description": "来自评测数据的真实增强款式图",
            "tags": ["训练集", "真实款式"],
            "image": f"/assets/styles/style-{sid:02d}.png",
            "source_url": row["enhanced_url"],
            "original_url": row["original_url"],
            "tip_count": tip_count,
        })
        print(f"style {sid:02d}: downloaded={ok}, generated_tips={tip_count}")

    MANIFEST_PATH.write_text(json.dumps({"styles": styles}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {MANIFEST_PATH}")


if __name__ == "__main__":
    main()
